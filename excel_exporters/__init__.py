import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
import sqlite3
from datetime import datetime
import os

class ExcelExporter:
    """Handles Excel export functionality with formatting similar to the templates"""
    
    def __init__(self, db_path="database/extraction_data.db"):
        self.db_path = db_path
    
    def create_annexure4_export(self, output_file='Annexure4_Formatted_Export.xlsx'):
        """Create formatted Excel export for Annexure IV data matching image 4 format"""
        
        # Get data from database
        conn = sqlite3.connect(self.db_path)
        
        # Query to get the summary format
        query = '''
            SELECT 
                ROW_NUMBER() OVER (ORDER BY a.id) as "File Count",
                a.pli_request_no as "PLI Request No",
                a.ifci_no as "IFCI No", 
                a.file_name as "File Name",
                a.date_processed as "Date",
                a.file_name as "Annexure IV - Declaration from PDF",
                a.applicant as "Applicant",
                a.supplier as "Supplier",
                a.subject as "Subject",
                a.sr_no_1 as "Sr No 1",
                a.sr_no_2 as "Sr No 2", 
                a.sr_no_3 as "Sr No 3",
                a.sr_no_4 as "Sr No 4",
                a.sr_no_5 as "Sr No 5",
                a.sr_no_6 as "Sr No 6",
                a.applicant_signatory_ca_fa as "Applicant signatory CA / SA",
                a.udin_values as "UDIN values",
                a.udin_values_2 as "UDIN values 2"
            FROM annexure4_data a
            ORDER BY a.id
        '''
        
        df = pd.read_sql_query(query, conn)
        
        # Get detailed parts data for content text
        parts_query = '''
            SELECT 
                a.id as annexure_id,
                GROUP_CONCAT(
                    COALESCE(p.part_no, '') || ' ' || 
                    COALESCE(p.part_description, '') || ' ' ||
                    COALESCE(p.selling_price_inr, '') || ' ' || 
                    COALESCE(p.value_direct_import_inr, '') || ' ' ||
                    COALESCE(p.broad_description_parts_imported, '') || ' ' || 
                    COALESCE(p.value_parts_imported_suppliers_inr, '') || ' ' ||
                    COALESCE(p.broad_description_parts_imported_suppliers, '') || ' ' || 
                    COALESCE(p.local_content, '') || ' ' || 
                    COALESCE(p.dva_percentage, ''),
                    ' | '
                ) as content_text
            FROM annexure4_data a
            LEFT JOIN annexure4_parts p ON a.id = p.annexure4_id
            GROUP BY a.id
        '''
        
        parts_df = pd.read_sql_query(parts_query, conn)
        conn.close()
        
        # Add content text column
        if not parts_df.empty:
            df['Content text matching exact as per output format'] = parts_df['content_text'].fillna('')
        else:
            df['Content text matching exact as per output format'] = ''
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Annexure IV"
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Add headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, max 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add notes at the bottom
        notes_row = len(df) + 3
        notes = [
            "sr no 2: Only match Table header and second row(A, B) and point content (not applicable for values)",
            "sr no 3: just match the title not values",
            "sr no 4: same as 2 and 3",
            "sr no 5: same as 2 and 4", 
            "sr no 6: same as 2 and 5"
        ]
        
        for i, note in enumerate(notes):
            ws.cell(row=notes_row + i, column=1, value=note)
            ws.cell(row=notes_row + i, column=1).font = Font(italic=True, size=9)
        
        # Save workbook
        wb.save(output_file)
        print(f"✓ Formatted Annexure IV export created: {output_file}")
        return output_file
    
    def create_comprehensive_report(self, output_file='Comprehensive_Data_Report.xlsx'):
        """Create a comprehensive report with all data"""
        
        conn = sqlite3.connect(self.db_path)
        
        # Get all data
        annexure4_query = '''
            SELECT 
                a.*,
                p.part_no, p.part_description, p.selling_price_inr,
                p.value_direct_import_inr, p.broad_description_parts_imported,
                p.value_parts_imported_suppliers_inr, p.broad_description_parts_imported_suppliers,
                p.local_content, p.dva_percentage
            FROM annexure4_data a
            LEFT JOIN annexure4_parts p ON a.id = p.annexure4_id
            ORDER BY a.id, p.id
        '''
        
        annexure6_query = 'SELECT * FROM annexure6_data ORDER BY created_at DESC'
        processing_log_query = 'SELECT * FROM processing_log ORDER BY created_at DESC'
        
        annexure4_df = pd.read_sql_query(annexure4_query, conn)
        annexure6_df = pd.read_sql_query(annexure6_query, conn)
        log_df = pd.read_sql_query(processing_log_query, conn)
        
        conn.close()
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # Annexure IV data
            if not annexure4_df.empty:
                annexure4_df.to_excel(writer, sheet_name='Annexure IV Details', index=False)
            
            # Annexure VI data  
            if not annexure6_df.empty:
                annexure6_df.to_excel(writer, sheet_name='Annexure VI Details', index=False)
            
            # Processing log
            if not log_df.empty:
                log_df.to_excel(writer, sheet_name='Processing Log', index=False)
            
            # Summary statistics
            summary_data = {
                'Metric': [
                    'Total Annexure IV Files',
                    'Total Annexure VI Files', 
                    'Total Parts/Components (A4)',
                    'Total Invoice Items (A6)',
                    'Report Generated'
                ],
                'Value': [
                    len(annexure4_df['file_name'].dropna().unique()) if not annexure4_df.empty else 0,
                    len(annexure6_df['file_name'].dropna().unique()) if not annexure6_df.empty else 0,
                    len(annexure4_df[annexure4_df['part_no'].notna()]) if not annexure4_df.empty else 0,
                    len(annexure6_df) if not annexure6_df.empty else 0,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        print(f"✓ Comprehensive report created: {output_file}")
        return output_file
    
    def export_template_format(self, template_type='annexure4', output_file=None):
        """Export data in the exact template format"""
        
        if output_file is None:
            output_file = f'{template_type}_template_format.xlsx'
        
        if template_type == 'annexure4':
            try:
                return self.create_annexure4_template_format(output_file)
            except Exception as e:
                print(f"❌ Error creating template format: {e}")
                return None
        else:
            print(f"Template type '{template_type}' not supported yet.")
            return None
    
    def create_annexure4_template_format(self, output_file):
        """Create Annexure IV export in exact template format"""
        
        conn = sqlite3.connect(self.db_path)
        
        # Get data in the exact format needed
        query = '''
            SELECT 
                a.file_name,
                a.date_processed,
                a.applicant,
                a.supplier,
                a.subject,
                a.currency_name,
                a.reference_date,
                a.foreign_exchange_rate,
                a.iec_code,
                p.*
            FROM annexure4_data a
            LEFT JOIN annexure4_parts p ON a.id = p.annexure4_id
            ORDER BY a.id, p.id
        '''
        
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        if df.empty:
            print("No data found to export.")
            return None
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Annexure IV Declaration"
        
        # Create header section (similar to PDF format)
        ws.merge_cells('A1:I1')
        ws['A1'] = "XYZ LIMITED"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A3:I3')
        ws['A3'] = "Annexure-IV: Format of Undertaking from Suppliers of the Application"
        ws['A3'].font = Font(bold=True, size=12)
        
        # Add date and other header info
        if not df.empty:
            first_row = df.iloc[0]
            ws['A5'] = f"Date: {first_row['date_processed']}"
            ws['A7'] = "To,"
            ws['A8'] = first_row['applicant'] if pd.notna(first_row['applicant']) else "ABC LIMITED"
            ws['A10'] = f"Sub: {first_row['subject']}" if pd.notna(first_row['subject']) else "Sub: Declaration of import in the parts / components / aggregates"
        
        # Create table headers
        table_start_row = 15
        headers = [
            "Part No.", "Part Description", "Selling Price in INR (excluding GST)",
            "Value of Direct Import by us (in INR)*", "Broad Description of parts imported by us",
            "Value of parts imported by our suppliers (in INR)*", 
            "Broad Description of parts imported by our suppliers", "Local Content", "DVA%"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add column labels (A), (B), (C), etc.
        label_row = table_start_row + 1
        labels = ['(A)', '(B)', '(C)', '(D)', '(E)', '(F)', '(G)', '(H)', '(I)']
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=label_row, column=col, value=label)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        
        # Add data rows
        data_start_row = table_start_row + 2
        current_row = data_start_row
        
        for _, row in df.iterrows():
            if pd.notna(row['part_no']):  # Only add rows with part numbers
                values = [
                    row['part_no'], row['part_description'], row['selling_price_inr'],
                    row['value_direct_import_inr'], row['broad_description_parts_imported'],
                    row['value_parts_imported_suppliers_inr'], row['broad_description_parts_imported_suppliers'],
                    row['local_content'], row['dva_percentage']
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col, value=value or '')
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                current_row += 1
        
        # Add footer notes
        footer_row = current_row + 2
        ws.merge_cells(f'A{footer_row}:I{footer_row}')
        ws[f'A{footer_row}'] = "* Import Value = CIF + Import Duty (but excluding GST). Imports will include royalty, license technical know-how expenses and imported services, in line with the declarations made under GST."
        ws[f'A{footer_row}'].font = Font(size=9)
        
        # Add IEC code
        if not df.empty and pd.notna(df.iloc[0]['iec_code']):
            ws[f'A{footer_row + 2}'] = f"Our IEC Code: {df.iloc[0]['iec_code']}"
        
        # Adjust column widths
        column_widths = [10, 25, 15, 20, 25, 20, 25, 12, 10]
        for i, width in enumerate(column_widths, 1):
            column_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[column_letter].width = width
        
        wb.save(output_file)
        print(f"✓ Template format export created: {output_file}")
        return output_file

# Helper function for easy usage
def export_all_data(db_path="database/extraction_data.db"):
    """Export all data in various formats"""
    exporter = ExcelExporter(db_path)
    
    files_created = []
    
    try:
        # Create formatted Annexure IV export
        file1 = exporter.create_annexure4_export()
        files_created.append(file1)
    except Exception as e:
        print(f"❌ Error creating Annexure IV export: {e}")
    
    try:
        # Create comprehensive report
        file2 = exporter.create_comprehensive_report()
        files_created.append(file2)
    except Exception as e:
        print(f"❌ Error creating comprehensive report: {e}")
    
    try:
        # Create template format
        file3 = exporter.export_template_format('annexure4')
        if file3:
            files_created.append(file3)
    except Exception as e:
        print(f"❌ Error creating template format: {e}")
    
    print(f"\n✓ Created {len(files_created)} export files:")
    for file in files_created:
        print(f"  - {file}")
    
    return files_created
